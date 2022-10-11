#coding:utf-8
import random
import csv
import io
import os
import configparser
import codecs
from distutils.util import strtobool
from traceback import format_exc
from flask import jsonify

def remove_prefix(text, prefix):
    if text.startswith(prefix):
        return text[len(prefix):]
    return text

def remove_suffix(text, suffix):
    if text.endswith(suffix):
        return text[:-len(suffix)]
    return text

def average(L):
    if len(L) == 0:
        return 0
    else:
        return sum(L)/len(L)

def count(L):
    l = [i for i in L if i != 0]
    return (len(l))

def sum_abs_diff(a, b):  # Compute sum of absolute values of a minus b
    if shape(a) != shape(b):
        return None
    s = 0
    if isinstance(a[0],list):
        for (rowa,rowb) in zip(a,b):
            s += sum(abs(x-y) for (x,y) in zip(rowa,rowb))
    else:
        s += sum(abs(x-y) for (x,y) in zip(a,b))
    return s

def shape(M):
    # Simply assume that all rows have equal length
    if not isinstance(M,list): return None
    if not isinstance(M[0],list): return len(M)
    return (len(M),len(M[0]))

#??????
def random_id(length=8):
    chars = '0123456789ABCDEFGHIJKLMNOPQRSTUVWXYZabcdefghijklmnopqrstuvwxyz'
    s = "".join(random.sample(chars, length))
    return s

def subtract_m(A, B):
    C = []
    for (a,b) in zip(A, B):
        c = [x - y for (x,y) in zip(a,b)]
        C.append(c)
    return C

def read_csv(filename):
    with io.open(filename, mode="r", newline='', encoding='utf-8') as f:
        for row in csv.reader(f, skipinitialspace=True):
            yield [cell for cell in row]

def read_xlsx(filename):
    import openpyxl
    book = openpyxl.load_workbook(filename)
    sheet = book.active
    for row in sheet.rows:
        yield [cell.value for cell in row]

# def load_constituencies(confile):
#     """Load constituencies from a file."""
#     if confile.endswith("csv"):
#         reader = read_csv(confile)
#     else:
#         reader = read_xlsx(confile)
#     cons = []
#     for row in reader:
#         try:
#             assert(int(row[1]) + int(row[2]) > 0)
#         except Exception as e:
#             print(row[1:3])
#             raise Exception("Error loading constituency file: "
#                             "constituency seats and adjustment seats "
#                             "must add to a nonzero number.")
#         cons.append({
#             "name": row[0],
#             "num_fixed_seats": int(row[1]),
#             "num_adj_seats": int(row[2])})
#     return cons

def isPosInt(x):
    if isinstance(x, float) and round(x) == x and x >= 0: return True
    if isinstance(x, str) and x.isdigit() and x.isascii(): return True
    if isinstance(x,int) and x >= 0: return True
    return False

def load_votes_from_excel(stream, filename):
    lines = []
    if not hasattr(stream, "seekable") and hasattr(stream, "_file"):
        stream.seekable = stream._file.seekable
        # TODO: Remove this monkeypatch once this issue has been resolved.
    import openpyxl
    if stream:
        book = openpyxl.load_workbook(stream)
    else:
        book = openpyxl.load_workbook(filename)
    sheet = book.active
    for row in sheet.rows:
        line = []
        for cell in row:
            value = round(cell.value, 10) if isinstance(cell.value, float) else cell.value
            line.append(value)
        lines.append(line)
    return lines

def remove_blank_rows(rows):
    for i in range(len(rows)-1, 0, -1):
        row = rows[i]
        if any(x is not None for x in row):
            break
        rows.pop()
    return rows

def correct_deprecated(L):
    # Remove 1-, 2- etc. from deprecated values in L["systems"]
    import re
    deprec_list = [
        "adj_alloc_divider",          "adj_determine_divider",
        "adjustment_allocation_rule", "adjustment_division_rule",
        "adjustment_method",          "constituency_allocation_rule",
        "name",                       "primary_divider"]
    old_names = {
        "norwegian-icelandic": "max-const-seat-share",
        "pure-vote-ratios":    "max-const-vote-percentage",
        "nearest-neighbor":    "nearest-to-previous",
        "nearest-to-last":     "nearest-to-previous",
    }
    translate = {
        "constituency_allocation_rule": "primary_divider",
        "adjustment_division_rule":     "adj_determine_divider",
        "adjustment_allocation_rule":   "adj_alloc_divider",
    }
    for sys in L["systems"]:
        for deprec in deprec_list:
            if deprec in sys:
                sys[deprec] = re.sub('^[0-9AB]-', '', sys[deprec])
        for (oldkey,newkey) in translate.items():
            if oldkey in sys:
                sys[newkey] = sys[oldkey]
        for (old,new) in old_names.items():
            if sys["adjustment_method"] == old:
                sys["adjustment_method"] = new
    return L

def add_empty_party_votes(vote_table):
    vote_table["party_vote_info"] = {
        "name": "–",
        "num_fixed_seats": 0,
        "num_adj_seats": 0,
        "votes": [],
        "specified": False,
        "total": 0
    }
    return vote_table;

def process_vote_table(rows, filename):

    # ERROR CHECKING
    row_lengths = [len(row) for row in rows]
    if min(row_lengths) < max(row_lengths):
        return 'Not all rows have equal length'
    if row_lengths[0] < 2:
        return 'Fewer than two columns'
    elif len(rows) < 2:
        return 'Only one row'
    toprow = rows[0]
    if  toprow[1].lower() not in ["fixed", "cons"]:
        return 'Heading of second column must be "fixed" (for fixed seats)'
    if  toprow[2].lower() != "adj":
        return 'Heading of third column must be "adj" (for adjustment seats)'
    if not all(toprow[3:]):
        return 'Some party names are blank'

    second_last_blank = all(x is None for x in rows[-2])
    num_const = len(rows) - (3 if second_last_blank else 1)
        
    if not all(l[0] for l in rows[1:num_const+1]):
        return 'Some constituency names are blank'
    if second_last_blank and rows[-1][0] is None:
        return 'The party vote name is blank'
    
    for row in rows[1:]:
        for i in range(3,len(row)):
            if row[i] is None:
                row[i] = 0
            elif isPosInt(row[i]):
                row[i] = int(row[i])
            else:
                return 'All votes must be non-negative integer numbers, e.g. no values obtained by formulas'

    # BUILD A DICTIONARY RES WITH ALL VOTING INFORMATION
    res = {}
    num_const = len(rows) - (3 if second_last_blank else 1)    
    res["votes"] = [[parsint(v) for v in row[3:]] for row in rows[1:num_const+1]]
    res["parties"] = rows[0][3:]

    res["constituencies"] = [{
        "name": row[0],
        "num_fixed_seats": parsint(row[1]),
        "num_adj_seats": parsint(row[2])
    } for row in rows[1:num_const+1]]

    res["name"] = determine_table_name(rows[0][0], filename)
    
    if second_last_blank:
        party_vote_info = rows[-1][3:]
        res["party_vote_info"] = {
            "name": rows[-1][0],
            "num_fixed_seats": rows[-1][1],
            "num_adj_seats": rows[-1][2],
            "votes": party_vote_info if any(party_vote_info) else [""]*len(party_vote_info),
            "specified": True,
            "total": sum(party_vote_info) if any(party_vote_info) else ""
        }
        if res["party_vote_info"]["num_fixed_seats"] is None:
            res["party_vote_info"]["num_fixed_seats"] = 0
        if res["party_vote_info"]["num_adj_seats"] is None:
            res["party_vote_info"]["num_adj_seats"] = 0
    else:
        res = add_empty_party_votes(res)
    return res

def parsint(value):
    return int(value) if value else 0

def determine_table_name(first,filename):
    return first if first is not None else os.path.basename(os.path.splitext(filename)[0])

def hms(sec):
    # Turn seconds into xxx days hh:mm:ss
    sec = round(sec)
    minute = sec//60
    sec %= 60
    hr = minute//60
    minute %= 60
    d = hr//24
    hr %= 24
    s = f"{minute:02}:{sec:02}"
    if hr > 0:
        s = f"{hr:02}" + s
    if d == 1:
        s = "1 day + " + s
    elif d > 1:
        s = f"{d} days + " + s
    return s

dispcompact=False

def disp(value=None, width=None, title=None):
    if width is None:
        width=disp.width
    else:
        disp.width=width
    if isinstance(value, (type({}.keys()),type({}.values()))):
        value = list(value)
    from pprint import PrettyPrinter
    pp = PrettyPrinter(compact=dispcompact, width=width).pprint
    if title:
        print("\n" + title.upper() + ":")
    if not value: return
    pp(value)
disp.width=100

def dispv(title, value=None, ndec=3):
    import numpy as np
    if value is None:
        value = title
        title=''
    if title:
        print("\n" + title.upper() + ":")
    np.set_printoptions(suppress=True, precision=ndec, floatmode="fixed",
                        linewidth=120)
    print(np.array(value))

def writecsv(file, L):
    # Writes list of lists L to csv-file
    import csv
    with open(file, "w") as f:
        writer = csv.writer(f)
        writer.writerows(L)

def get_cpu_count():
    from multiprocessing import cpu_count
    return cpu_count()
        
def get_cpu_counts():
    from math import sqrt
    cpu_counts = []
    rcount = sqrt(2)
    count = round(rcount)
    cpu_count = get_cpu_count()
    while count <= cpu_count:
        cpu_counts.append(count)
        rcount *= sqrt(2)
        count = round(rcount)
    return cpu_counts

def timestamp():
    from datetime import datetime as dt
    ts = dt.now().strftime("%H:%M:%S.%f")[:-3]
    return ts

def timestampmsg(s):
    print(timestamp() + ": " + s)
